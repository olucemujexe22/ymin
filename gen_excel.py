# -*- coding: utf-8 -*-
"""应用中心数据采集表 — 以模块为单位，每个模块可填多条规格"""
import openpyxl, os, subprocess, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 预解析 JS 数据为 JSON =====
if not os.path.exists('_specs.json'):
    subprocess.run(['node', '-e', 
        "var fs=require('fs');var c=fs.readFileSync('data/system-modules.js','utf8');"
        "var m=c.match(/var SPECS = ({[\\s\\S]*?});\\s*\\n\\s*function getSpecs/);"
        "if(m){var s=eval('('+m[1]+')');fs.writeFileSync('_specs.json',JSON.stringify(s,null,2))}"])
if not os.path.exists('_db.json'):
    subprocess.run(['node', '-e',
        "var fs=require('fs');var c=fs.readFileSync('data/system-modules.js','utf8');"
        "var i=c.indexOf('var DB =');var e=c.indexOf('}; // end DB');"
        "if(i>=0&&e>=0){var s=eval('('+c.substring(i+8,e+1)+')');fs.writeFileSync('_db.json',JSON.stringify(s,null,2))}"])

with open('_specs.json', 'r', encoding='utf-8') as f:
    SPECS = json.load(f)
with open('_db.json', 'r', encoding='utf-8') as f:
    DB = json.load(f)

# ===== 样式 =====
wb = openpyxl.Workbook()
hdrF = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
hdrFill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
nF = Font(name='微软雅黑', size=9); bF = Font(name='微软雅黑', size=9, bold=True)
yel = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
grn = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
gry = PatternFill(start_color='EDEFF5', end_color='EDEFF5', fill_type='solid')
pur = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
bdr = Border(left=Side(style='thin',color='00CCCCCC'),right=Side(style='thin',color='00CCCCCC'),
             top=Side(style='thin',color='00CCCCCC'),bottom=Side(style='thin',color='00CCCCCC'))
wT = Alignment(wrap_text=True, vertical='top'); wC = Alignment(wrap_text=True, vertical='center', horizontal='center')

def sc(ws, r, c, v, font=None, fill=None, align=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = font or nF; cl.border = bdr; cl.alignment = align or wT
    if fill: cl.fill = fill
    return cl

def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ===== 数据查找 =====
def find_db_mods(pk, tk, sn):
    """从DB查找模块，支持模糊匹配"""
    try:
        td = DB.get(pk, {}).get(tk, {})
    except: return None
    if sn in td: return td[sn]
    for k, v in td.items():
        if sn in k or k in sn: return v
    return list(td.values())[0] if td else None

def find_specs(pk, tk, sn):
    """从SPECS查找规格，支持模糊匹配"""
    try:
        td = SPECS.get(pk, {}).get(tk, {})
    except: return []
    if sn in td: return td[sn]
    for k, v in td.items():
        if sn in k or k in sn: return v
    return []

DEF_MODS = [
    {'name':'输入滤波','desc':'电源输入端需滤波电容以抑制纹波与噪声，保证后续电路稳定工作。'},
    {'name':'储能/旁路','desc':'储能与旁路电容提供瞬态电流支持，减小电源总线波动。'},
    {'name':'输出滤波','desc':'输出端需滤波电容以平滑电压波形，确保负载端获得纯净电源。'},
    {'name':'保护/去耦','desc':'去耦与保护电容吸收高频噪声和电压尖峰，保障系统EMC合规与长期可靠性。'},
]

PAGES = {
    '汽车电子': {'pk':'automotive','tabs':[
        {'k':'t0','tk':'电机驱动','nm':'电机驱动','sa':['电机控制器-MCU','直流转直流-DCDC','高压分线盒-PDU','整车控制-VCU','电池管理系统-BMS']},
        {'k':'t1','tk':'安全部件','nm':'安全部件','sa':['刹车制动-BS','电子助力转向-EPS','安全气囊','车窗控制器','座椅通风','车身稳定系统-ESC']},
        {'k':'t2','tk':'热管理部件','nm':'热管理部件','sa':['电子水泵/油泵','空调压缩机控制器','BMS电池管理','冷却风扇控制器','电子水阀','PTC加热器','电子水泵']},
        {'k':'t3','tk':'lamp','nm':'新能源车车灯','sa':['新能源车车灯']},
        {'k':'t4','tk':'adas','nm':'智能驾驶','sa':['底盘域控制器','雷达','域控制器-DCU','雷达摄像头','智能座舱','车门窗控制器','自动换挡器','远程通信模块/T-BOX']},
        {'k':'t5','tk':'cockpit','nm':'多媒体系统','sa':['组合仪表、车载屏、导航','抬头显示HUD','远程通讯模块-T-BOX','音响']},
        {'k':'t6','tk':'aux2','nm':'辅助','sa':['冰箱控制器','USB/WPT']},
        {'k':'t7','tk':'obc','nm':'车载OBC','sa':['OBC-DCDC']},
        {'k':'t8','tk':'charger','nm':'充电桩','sa':['交流转直流-AC/DC','直流转直流-DC-DC']},
    ]},
    'AI服务器与数据中心': {'pk':'aiserver','tabs':[
        {'k':'t0','tk':'cpu','nm':'AI服务器主板&显卡','sa':['图形处理器-GPU、中央处理器 CPU','DC-DC转换器']},
        {'k':'t1','tk':'aux','nm':'AI服务器电源','sa':['主供电源-低压输出端','主供电源-高压输入端','BBU备用电源']},
        {'k':'t2','tk':'mem','nm':'AI服务器存储','sa':['SSD企业级固态硬盘','数据存储盘']},
    ]},
    '储能': {'pk':'energy','tabs':[
        {'k':'t0','tk':'pv','nm':'逆变器','sa':['逆变器','定日镜']},
        {'k':'t1','tk':'pcs','nm':'变流器','sa':['变流器']},
        {'k':'t2','tk':'bms','nm':'BMS','sa':['电池管理系统-BMS']},
    ]},
    '新型电机驱动': {'pk':'motor','tabs':[
        {'k':'t0','tk':'inverter','nm':'智慧出行','sa':['智慧出行']},
        {'k':'t1','tk':'servo','nm':'机器人','sa':['机器人']},
        {'k':'t2','tk':'stepper','nm':'无人机','sa':['无人机']},
        {'k':'t3','tk':'bldc','nm':'电动工具','sa':['电动工具']},
        {'k':'t4','tk':'thermal2','nm':'汽车热管理','sa':['汽车热管理']},
        {'k':'t5','tk':'home','nm':'智能家电','sa':['智能家电']},
    ]},
    '仪器仪表': {'pk':'instrument','tabs':[
        {'k':'t0','tk':'precision','nm':'智能电表','sa':['智能电表','载波模块','集中器','新融合终端','DTU (配电自动化终端)','开关电源','断路器']},
        {'k':'t1','tk':'industrial','nm':'智能燃气表','sa':['智能燃气表']},
        {'k':'t2','tk':'medical','nm':'智能水表','sa':['智能水表、热量表']},
    ]},
    '消费类电子': {'pk':'consumer','tabs':[
        {'k':'t0','tk':'smartphone','nm':'PD快充','sa':['快充输入端专用','快充输出端专用']},
        {'k':'t1','tk':'led','nm':'智能照明','sa':['智能照明']},
        {'k':'t2','tk':'pen','nm':'电子笔','sa':['电子笔']},
        {'k':'t3','tk':'digital','nm':'数码3C','sa':['数码3C']},
        {'k':'t4','tk':'security','nm':'安防','sa':['安防']},
    ]},
}

HEADERS = [
    '一级Tab','子应用名称','子应用描述','推荐产品系列','电路拓扑图',
    '模块名称','模块描述',
    '规格-系列','规格-料号','电压','容量','尺寸','ESR','纹波电流','寿命','备注(PDF来源)'
]
COLS = len(HEADERS)

# ===== Sheet 1: 使用说明 =====
ws0 = wb.active; ws0.title = '使用说明'
ws0.merge_cells('A1:F1')
ws0.cell(row=1, column=1, value='应用中心数据采集表 — 使用说明').font = Font(name='微软雅黑', bold=True, size=14)
for i, n in enumerate([
    '1. 每个"模块"占一行，其推荐规格紧接下方。一个子应用有4个模块，每个模块可填多条规格。',
    '2. 颜色：⬜灰=Tab标题 | 🟣紫=模块行 | 🟡黄=待事业部填写 | 🟢绿=已有AI演示参考数据',
    '3. "子应用描述"：一句话功能描述。显示在详情页标题下方。',
    '4. "推荐产品系列"：逗号分隔，如 VHT,NPX,CW3H。',
    '5. "电路拓扑图"：填写拓扑图文件名，或标注"使用SVG方块图"。',
    '6. "模块名称/描述"：对应详情页"系统模块详解"卡片。🟢绿=已从应用手册预填，请核对。',
    '7. "推荐规格"：每行一个系列/料号。🟢绿=已从应用手册预填。🟡黄=需全新填写。',
    '8. 填完后发回数字智能部，用于批量替换代码中的占位数据。',
]):
    ws0.cell(row=3+i, column=1, value=n).font = Font(name='微软雅黑', size=10)
cw(ws0, [80, 15, 15, 15, 15, 15])

# ===== 每个领域 Sheet =====
for pn, pd in PAGES.items():
    ws = wb.create_sheet(pn); pk = pd['pk']; row = 2
    for c, h in enumerate(HEADERS, 1):
        sc(ws, 1, c, h, font=hdrF, fill=hdrFill, align=wC)
    
    for tab in pd['tabs']:
        # Tab标题行
        sc(ws, row, 1, f"▸ {tab['nm']}（{len(tab['sa'])}个子应用）", font=bF, fill=gry)
        for c in range(2, COLS+1): sc(ws, row, c, '', fill=gry)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
        row += 1
        
        for sn in tab['sa']:
            mods = find_db_mods(pk, tab['tk'], sn)
            if not mods: mods = DEF_MODS
            sps = find_specs(pk, tab['tk'], sn)
            
            for mi, mod in enumerate(mods):
                is_first = (mi == 0)
                rdat = ([tab['nm'], sn, '', '', ''] if is_first else ['','','','','']) + [mod['name'], mod['desc']] + ['']*9
                for c, v in enumerate(rdat, 1):
                    fl = yel if is_first and c <= 5 else (pur if c in [6,7] else (grn if sps and c >= 8 else yel))
                    sc(ws, row, c, v, fill=fl)
                
                if sps and mi == 0:
                    row += 1  # 规格从模块下一行开始
                    for si, sp in enumerate(sps):
                        if si > 0: row += 1
                        sd = ['']*7 + [sp.get('series',''), sp.get('pn',''), sp.get('voltage',''),
                              sp.get('cap',''), sp.get('size',''), sp.get('esr',''),
                              sp.get('ripple',''), sp.get('life',''), sp.get('note','')]
                        for c, v in enumerate(sd, 1):
                            sc(ws, row, c, v, fill=grn if c >= 8 else None)
                row += 1
            row += 1
    
    cw(ws, [14, 26, 28, 18, 18, 16, 32, 10, 24, 8, 10, 12, 10, 10, 16, 38])
    ws.freeze_panes = 'A2'

# ===== 总览页 =====
ws_ov = wb.create_sheet('总览页')
oh = ['领域', '简介文案（当前）', '简介文案（修改为）', '特性标签（当前）', '特性标签（修改为）', '推荐系列（当前）', '推荐系列（修改为）']
for c, h in enumerate(oh, 1): sc(ws_ov, 1, c, h, font=hdrF, fill=hdrFill, align=wC)
for r, rd in enumerate([
    ['汽车电子', '主驱/电驱电控、充电系统、安全部件、ADAS、热管理、智能座舱、车灯、充电桩 —— AEC-Q200合规，耐高温135℃，抗振动。', '', 'AEC-Q200 / 135℃耐高温 / 极低漏电流', '', '8个应用子领域 · 10+产品系列', ''],
    ['AI服务器', 'GPU 加速卡、CPU 供电、PDN 总线架构的大电流、超低 ESR 电容方案，支持 400V/800V DC-Link。', '', '大电流 / 超低ESR / 小型化', '', 'VHT · NPX · CW3H系列', ''],
    ['仪器仪表', '精密测量仪器、工业自动化仪表、医疗检测设备等对精度和稳定性要求极高的电容方案。', '', '高精度 / 低漏电流 / 长寿命', '', 'VKM · VPT · LK系列', ''],
    ['新型电机驱动', '变频器、伺服驱动器、步进电机驱动等工业电机控制系统的 DC-Link 与滤波电容方案。', '', 'DC-Link / 滤波 / 高纹波', '', 'VPG · VPX · CW3H系列', ''],
    ['储能', '光伏逆变器、储能系统、充电桩等高压大容量电容方案，支持高纹波电流与长寿命需求。', '', '高压大容量 / 高纹波 / 长寿命', '', 'CW3H · VPG系列', ''],
    ['消费类电子', '笔记本电脑、智能家居、PD快充、LED照明等消费电子产品的薄型化小型化电容方案。', '', '薄型化 / 小型化 / 低ESR', '', 'NPX · VKM · VPT系列', ''],
], 2):
    for c, v in enumerate(rd, 1):
        sc(ws_ov, r, c, v, fill=yel if c in [3,5,7] else None)
cw(ws_ov, [16, 42, 42, 26, 26, 22, 22])

# ===== 数据字典 =====
ws_dd = wb.create_sheet('数据字典')
for c, h in enumerate(['字段名','对应代码位置','说明','示例'], 1):
    sc(ws_dd, 1, c, h, font=hdrF, fill=hdrFill, align=wC)
for r, rd in enumerate([
    ['子应用描述','appData[tab].description','一句话功能描述','车载高压电池输入端滤波，需承受高压与大纹波电流'],
    ['推荐产品系列','appData[tab].subApps[i].series','逗号分隔','VHT,NPX,CW3H'],
    ['电路拓扑图','<img src="...">','文件名','topo_automotive_mcu.png'],
    ['模块名称','DB[page][tab][subApp][i].name','系统功能模块名','高压输入滤波'],
    ['模块描述','DB[...][i].desc','模块功能详述（含推荐系列）','车载高压电池输入端滤波...推荐VHT系列，ESR≤8mΩ'],
    ['规格-系列','SPECS[...][i].series','产品系列代号','VHT'],
    ['规格-料号','SPECS[...][i].pn','具体料号','VHTE1051V331MVC'],
    ['规格-电压/容量/尺寸等','SPECS[...][i]','额定参数','35V / 330µF / 10×10.5'],
    ['规格-备注','SPECS[...][i].note','数据来源（PDF页码+竞品对标）','[固液]P06 · Nichicon GYA/GYE'],
], 2):
    for c, v in enumerate(rd, 1): sc(ws_dd, r, c, v)
cw(ws_dd, [18, 40, 40, 50])

out = '应用中心数据采集表.xlsx'
wb.save(out)
print(f'✅ 已生成: {out}')
print(f'   Sheets: 使用说明 + 总览页 + 数据字典 + {len(PAGES)}个领域')

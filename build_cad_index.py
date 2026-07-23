"""从最新 3DCAD 文件夹生成网页索引和 Excel 清单。

运行方式::

    python build_cad_index.py

网页只发布标准 STEP 文件；Creo/ProE 的 .asm.N、.prt.N 与导出日志会单独
列入清单，但不会进入浏览器预览索引。所有重复项只做标记，不删除源文件。
"""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
CAD_ROOT = ROOT / "3dcad文件"
OUTPUT = ROOT / "data" / "cad-models.json"
JS_OUTPUT = ROOT / "data" / "cad-models.js"
GEOMETRY_CACHE = ROOT / "data" / "cad-geometry.json"
PRODUCT_MAP_OUTPUT = ROOT / "data" / "cad-product-map.json"
PRODUCT_COVERAGE_OUTPUT = ROOT / "data" / "cad-product-coverage.json"

DIMENSION_PATTERN = re.compile(
    r"(?<![\d.])(\d+(?:\.\d+)?)\s*[xX×*]\s*(\d+(?:\.\d+)?)"
    r"(?:\s*[xX×*]\s*(\d+(?:\.\d+)?))?(?![\d.])"
)
HYPHEN_DIMENSION_PATTERN = re.compile(
    r"(?<![\d.])(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)(?![\d.])"
)
COLLECTION_PATTERN = re.compile(r"^(\d{2})\s*(.+?)3dcad集合$", re.IGNORECASE)
VERSION_PATTERN = re.compile(r"^(\d{8})-3dcad$", re.IGNORECASE)

CATEGORY_NAMES = {
    "液态": "液态铝电解",
    "固态": "固态铝电解",
    "超电": "超级电容",
    "液态大型": "大型液态铝电解",
    "叠层": "叠层电容",
}
CATEGORY_ORDER = {
    "液态铝电解": 1,
    "固态铝电解": 2,
    "超级电容": 3,
    "大型液态铝电解": 4,
    "叠层电容": 5,
    "未分类": 99,
}
PACKAGE_ORDER = {
    "radial": 1,
    "smd": 2,
    "snapin": 3,
    "screw": 4,
    "t": 5,
    "stacked": 6,
    "other": 99,
}


def tidy_number(value: float | None) -> float | None:
    return None if value is None else round(float(value), 3)


def number_text(value: float) -> str:
    return f"{value:.3f}".rstrip("0").rstrip(".")


def format_dimensions(values: tuple[float, ...] | list[float] | None) -> str:
    if not values:
        return "—"
    return " × ".join(number_text(value) for value in values) + " mm"


def format_cad_envelope(values: tuple[float, ...] | list[float] | None) -> str:
    if not values or len(values) < 3:
        return "—"
    return " × ".join(f"{axis} {number_text(value)}" for axis, value in zip("XYZ", values[:3])) + " mm"


def normalize_dimension(value: Any) -> str:
    text = str(value or "").lower().replace("×", "x").replace("*", "x")
    return re.sub(r"[^0-9.x]", "", text)


def parse_dimensions(value: str) -> tuple[float, ...] | None:
    matches = list(DIMENSION_PATTERN.finditer(value))
    if matches:
        match = matches[-1]
        values = [float(match.group(1)), float(match.group(2))]
        if match.group(3) is not None:
            values.append(float(match.group(3)))
        return tuple(values)

    # “30-70 T型”一类文件名使用连字符；过滤产品编码中的大数字组合。
    matches = list(HYPHEN_DIMENSION_PATTERN.finditer(value))
    for match in reversed(matches):
        first, second = float(match.group(1)), float(match.group(2))
        if 1 <= first <= 200 and 1 <= second <= 500:
            return first, second
    return None


def find_nominal_dimensions(relative_path: Path) -> tuple[float, ...] | None:
    for part in reversed(relative_path.parts[:-1]):
        dimensions = parse_dimensions(part)
        if dimensions:
            return dimensions
    return None


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def unique(values: Iterable[Any]) -> list[Any]:
    result: list[Any] = []
    for value in values:
        if value in (None, "") or value in result:
            continue
        result.append(value)
    return result


def as_number(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_product_catalog() -> list[dict[str, Any]]:
    """读取原固态 CAD 制作表，用于保留系列、品号和脚距关联。"""
    candidates = [
        path
        for path in ROOT.parent.rglob("*.xlsx")
        if path.name.lower() == "固态-3dcad制作.xlsx".lower()
    ]
    if not candidates:
        return []
    book = load_workbook(candidates[0], data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_index, sheet in enumerate(book.worksheets[:2]):
        package_key = "radial" if sheet_index == 0 else "smd"
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if len(values) < 7 or not values[0]:
                continue
            diameter = as_number(values[3])
            length_index = 11 if package_key == "smd" and len(values) > 11 else 4
            length = as_number(values[length_index])
            if diameter is None or length is None:
                continue
            rows.append(
                {
                    "packageKey": package_key,
                    "itemNo": str(values[0]).strip(),
                    "series": str(values[1] or "").strip(),
                    "nominalKey": normalize_dimension(values[2]),
                    "diameter": diameter,
                    "length": length,
                    "pitch": as_number(values[6]),
                }
            )
    return rows


def closest_catalog_rows(
    candidates: Iterable[dict[str, Any]], diameter: float, length: float
) -> tuple[list[dict[str, Any]], float]:
    candidates = list(candidates)
    if not candidates:
        return [], float("inf")
    distances = [
        abs(row["diameter"] - diameter) + abs(row["length"] - length)
        for row in candidates
    ]
    minimum = min(distances)
    return [
        row
        for row, distance in zip(candidates, distances)
        if abs(distance - minimum) < 0.001
    ], minimum


def match_product_catalog(
    rows: list[dict[str, Any]],
    package_key: str,
    nominal_key: str,
    dimensions: tuple[float, ...] | None,
) -> tuple[list[dict[str, Any]], str]:
    if not dimensions or len(dimensions) < 2:
        return [], "unmatched"
    diameter, length = dimensions[:2]
    same_nominal = [
        row
        for row in rows
        if row["packageKey"] == package_key and row["nominalKey"] == nominal_key
    ]
    exact = [
        row
        for row in same_nominal
        if abs(row["diameter"] - diameter) <= 0.011
        and abs(row["length"] - length) <= 0.011
    ]
    if exact:
        return exact, "exact"
    nearest, distance = closest_catalog_rows(same_nominal, diameter, length)
    if nearest and distance <= 0.26:
        return nearest, "approximate"
    same_geometry = [
        row
        for row in rows
        if row["packageKey"] == package_key
        and abs(row["diameter"] - diameter) <= 0.011
        and abs(row["length"] - length) <= 0.011
    ]
    if same_geometry:
        return same_geometry, "geometry"
    return [], "unmatched"


def library_version(paths: Iterable[Path]) -> str:
    versions = [part for path in paths for part in path.parts if VERSION_PATTERN.match(part)]
    return max((VERSION_PATTERN.match(part).group(1) for part in versions), default="未标注")


def collection_info(relative_path: Path) -> tuple[str, str, int]:
    for part in relative_path.parts:
        match = COLLECTION_PATTERN.match(part)
        if not match:
            continue
        source_name = match.group(2).strip()
        return CATEGORY_NAMES.get(source_name, source_name), part, int(match.group(1))
    return "未分类", "未分类", 99


def package_info(relative_path: Path, category: str) -> tuple[str, str, str]:
    path_text = "/".join(relative_path.parts)
    if "贴片型" in path_text:
        subtype = "抗震座板" if "抗震" in path_text else ("普通座板" if "普通座板" in path_text else "")
        return "贴片型 SMD", "smd", subtype
    if "引线型" in path_text:
        subtype = "M脚" if "M脚" in path_text else ("直角折弯" if "直角折弯" in path_text else "")
        return "引线型 Radial", "radial", subtype
    if "牛角型" in path_text:
        subtype = "二脚" if "二脚" in path_text else ""
        return "牛角型 Snap-in", "snapin", subtype
    if "螺栓型" in path_text:
        return "螺栓型 Screw-terminal", "screw", ""
    if "T型" in path_text:
        return "T型", "t", ""
    if category == "叠层电容":
        return "叠层型", "stacked", ""
    return "其他", "other", ""


def load_existing_metadata() -> dict[tuple[str, str, str], dict[str, Any]]:
    """保留上一版固态库已匹配到的系列、品号和引脚间距。"""
    if not OUTPUT.exists():
        return {}
    try:
        existing = json.loads(OUTPUT.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    merged: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in existing:
        hash_value = row.get("fileHash")
        source = ROOT / str(row.get("step", ""))
        if not hash_value and source.is_file():
            hash_value = file_sha256(source)
        if not hash_value:
            continue
        key = (
            hash_value,
            str(row.get("packageKey", "")),
            normalize_dimension(row.get("nominal", "")),
        )
        target = merged.setdefault(key, {"sourceSeries": [], "sourceRefs": [], "pitches": [], "match": "unmatched"})
        legacy_series = list(row.get("sourceSeries") or row.get("series") or [])
        legacy_refs = list(row.get("sourceRefs") or [])
        if not legacy_refs:
            legacy_refs = [value for value in list(row.get("itemNos") or []) if re.fullmatch(r"F\d+", str(value), re.IGNORECASE)]
        target["sourceSeries"] = sorted(unique(target["sourceSeries"] + legacy_series))
        target["sourceRefs"] = sorted(unique(target["sourceRefs"] + legacy_refs))
        target["pitches"] = sorted(unique(target["pitches"] + list(row.get("pitches") or [])))
        if row.get("match") and row.get("match") != "unmatched":
            target["match"] = row["match"]
    return merged


def load_geometry_cache() -> dict[str, dict[str, Any]]:
    if not GEOMETRY_CACHE.exists():
        return {}
    try:
        payload = json.loads(GEOMETRY_CACHE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def dimensional_mismatch(
    nominal: tuple[float, ...] | None,
    actual: tuple[float, ...] | None,
    package_key: str,
) -> bool:
    if not nominal or not actual or len(nominal) < 2 or len(actual) < 2 or package_key == "stacked":
        return False
    return abs(nominal[0] - actual[0]) > 0.75 or abs(nominal[1] - actual[1]) > 1.5


def build_models() -> tuple[list[dict[str, Any]], list[dict[str, str]], str]:
    if not CAD_ROOT.is_dir():
        raise FileNotFoundError(f"未找到最新 CAD 目录：{CAD_ROOT}")

    step_files = sorted(CAD_ROOT.rglob("*.stp"), key=lambda path: path.as_posix().lower())
    if not step_files:
        raise FileNotFoundError(f"{CAD_ROOT} 中没有 STEP 文件")

    previous = load_existing_metadata()
    catalog_rows = load_product_catalog()
    geometry_cache = load_geometry_cache()
    models: list[dict[str, Any]] = []
    by_hash: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for step_file in step_files:
        relative_to_library = step_file.relative_to(CAD_ROOT)
        relative_to_root = step_file.relative_to(ROOT).as_posix()
        category, collection, collection_order = collection_info(relative_to_library)
        package, package_key, subtype = package_info(relative_to_library, category)
        nominal_dimensions = find_nominal_dimensions(relative_to_library)
        actual_dimensions = parse_dimensions(step_file.stem)
        working_dimensions = actual_dimensions or nominal_dimensions
        nominal_text = format_dimensions(nominal_dimensions or actual_dimensions)
        model_text = format_dimensions(working_dimensions)
        hash_value = file_sha256(step_file)
        geometry = geometry_cache.get(hash_value, {})
        cad_envelope = list(geometry.get("envelope") or []) if geometry.get("status") == "parsed" else []
        sorted_envelope = sorted(cad_envelope)
        metadata_key = (hash_value, package_key, normalize_dimension(nominal_text))
        metadata = previous.get(metadata_key, {}) if category == "固态铝电解" else {}
        matched_rows: list[dict[str, Any]] = []
        match_quality = str(metadata.get("match", "unmatched"))
        if category == "固态铝电解":
            matched_rows, catalog_quality = match_product_catalog(
                catalog_rows,
                package_key,
                normalize_dimension(nominal_text),
                actual_dimensions or nominal_dimensions,
            )
            if catalog_quality != "unmatched":
                match_quality = catalog_quality
        source_series = sorted(unique(list(metadata.get("sourceSeries") or []) + [row["series"] for row in matched_rows]))
        source_refs = sorted(unique(list(metadata.get("sourceRefs") or []) + [row["itemNo"] for row in matched_rows]))
        pitches = sorted(unique(list(metadata.get("pitches") or []) + [tidy_number(row["pitch"]) for row in matched_rows]))

        diameter = None
        length = None
        width = None
        depth = None
        height = None
        if working_dimensions:
            if package_key == "stacked" and len(working_dimensions) >= 3:
                width, depth, height = working_dimensions[:3]
                diameter, length = width, height
            else:
                diameter = working_dimensions[0]
                length = working_dimensions[1] if len(working_dimensions) > 1 else None

        path_hash = hashlib.sha1(relative_to_root.encode("utf-8")).hexdigest()[:10]
        model = {
            "id": f"cad-{path_hash}",
            "model": model_text,
            "nominal": nominal_text,
            "nominalDimensions": [tidy_number(value) for value in (nominal_dimensions or actual_dimensions or ())],
            "actualSize": format_dimensions(actual_dimensions),
            "cadEnvelope": [tidy_number(value) for value in cad_envelope],
            "cadEnvelopeText": format_cad_envelope(cad_envelope),
            "cadMin": tidy_number(sorted_envelope[0]) if sorted_envelope else None,
            "cadMax": tidy_number(sorted_envelope[-1]) if sorted_envelope else None,
            "cadParts": geometry.get("parts"),
            "cadVertices": geometry.get("vertices"),
            "cadTriangles": geometry.get("triangles"),
            "geometryStatus": geometry.get("status", "missing"),
            "fileName": step_file.name,
            "category": category,
            "collection": collection,
            "package": package,
            "packageKey": package_key,
            "subtype": subtype,
            "dimensions": [tidy_number(value) for value in (working_dimensions or ())],
            "diameter": tidy_number(diameter),
            "length": tidy_number(length),
            "width": tidy_number(width),
            "depth": tidy_number(depth),
            "height": tidy_number(height),
            "pitches": pitches,
            "sourceSeries": source_series,
            "sourceRefs": source_refs,
            "series": [],
            "itemNos": [],
            "productCandidateCount": 0,
            "productAmbiguousCount": 0,
            "step": relative_to_root,
            "fileSize": step_file.stat().st_size,
            "modified": datetime.fromtimestamp(step_file.stat().st_mtime).isoformat(timespec="seconds"),
            "fileHash": hash_value,
            "hashShort": hash_value[:12],
            "match": match_quality,
            "dimensionMismatch": dimensional_mismatch(nominal_dimensions, actual_dimensions, package_key),
            "_collectionOrder": collection_order,
            "_absolutePath": str(step_file),
        }
        models.append(model)
        by_hash[hash_value].append(model)

    duplicate_groups = [group for group in by_hash.values() if len(group) > 1]
    duplicate_groups.sort(key=lambda group: group[0]["step"])
    for index, group in enumerate(duplicate_groups, 1):
        group_id = f"DUP-{index:03d}"
        first_path = group[0]["step"]
        for model in group:
            model["duplicateGroup"] = group_id
            model["duplicateCount"] = len(group)
            model["duplicateOf"] = first_path if model["step"] != first_path else ""

    issues: list[dict[str, str]] = []
    for group in duplicate_groups:
        nominal_values = sorted({model["nominal"] for model in group})
        cross_size = len(nominal_values) > 1
        for model in group:
            issues.append(
                {
                    "level": "需核对" if cross_size else "提示",
                    "type": "完全重复（跨尺寸目录）" if cross_size else "完全重复",
                    "group": model["duplicateGroup"],
                    "category": model["category"],
                    "package": model["package"],
                    "nominal": model["nominal"],
                    "actual": model["actualSize"],
                    "file": model["fileName"],
                    "path": model["step"],
                    "description": f"与本组其他 {model['duplicateCount'] - 1} 个文件的 SHA-256 完全一致",
                    "suggestion": "核对目录归属；确认后可保留一个公共模型" if cross_size else "如无型号差异，可考虑仅保留一份",
                }
            )

    for model in models:
        if not model["dimensionMismatch"]:
            continue
        issues.append(
            {
                "level": "需核对",
                "type": "目录尺寸与文件名尺寸差异较大",
                "group": model.get("duplicateGroup", ""),
                "category": model["category"],
                "package": model["package"],
                "nominal": model["nominal"],
                "actual": model["actualSize"],
                "file": model["fileName"],
                "path": model["step"],
                "description": "直径差大于 0.75 mm，或高度差大于 1.5 mm",
                "suggestion": "核对 STEP 是否位于正确的名义尺寸目录",
            }
        )

    models.sort(
        key=lambda model: (
            model["_collectionOrder"],
            PACKAGE_ORDER.get(model["packageKey"], 99),
            model["diameter"] or 0,
            model["length"] or 0,
            model["step"],
        )
    )
    return models, issues, library_version(step_files)


def product_number(value: Any) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", str(value or "").strip())
    return tidy_number(float(match.group(0))) if match else None


def dimension_key(values: Iterable[float | None]) -> tuple[float, ...] | None:
    result = [value for value in values]
    if not result or any(value is None for value in result):
        return None
    return tuple(round(float(value), 2) for value in result if value is not None)


def match_token(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def load_production_products() -> list[dict[str, Any]]:
    candidates = sorted((ROOT / "data").glob("tp_good_full_*.json"))
    if not candidates:
        return []
    rows = json.loads(candidates[-1].read_text(encoding="utf-8"))
    return [row for row in rows if row.get("delete_time") is None]


def build_product_map(
    models: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    liquid = "液态铝电解"
    solid = "固态铝电解"
    supercapacitor = "超级电容"
    large_liquid = "大型液态铝电解"
    stacked = "叠层电容"
    column_categories = {
        82: liquid,
        592: liquid,
        83: solid,
        192: solid,
        587: solid,
        398: supercapacitor,
        589: supercapacitor,
        595: supercapacitor,
        84: stacked,
        594: stacked,
    }
    category_titles = {
        82: "液态铝电解电容器",
        592: "新型液态铝电解电容",
        83: "高分子固态铝电解电容器",
        192: "高分子混合动力铝电解电容器",
        587: "新型固液混合电容",
        398: "混合型超级电容（锂离子电容）",
        589: "新型超级电容",
        595: "新型超级电容",
        84: "叠层高分子固态铝电解电容器",
        594: "新型固态叠层电容",
    }
    package_by_shape = {
        "引线型": "radial",
        "贴片型": "smd",
        "牛角型": "snapin",
        "螺栓型": "screw",
        "基板自立型": "t",
    }

    model_groups: dict[tuple[str, str, tuple[float, ...]], list[dict[str, Any]]] = defaultdict(list)
    for model in models:
        dimensions = dimension_key(model.get("nominalDimensions") or [])
        if dimensions:
            model_groups[(model["category"], model["packageKey"], dimensions)].append(model)

    product_rows = load_production_products()
    mappings: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    counters: Counter[str] = Counter()
    model_items: dict[str, set[str]] = defaultdict(set)
    model_series: dict[str, set[str]] = defaultdict(set)
    model_ambiguous: dict[str, set[str]] = defaultdict(set)

    for product in product_rows:
        column_id = int(product.get("column_id") or 0)
        category = column_categories.get(column_id)
        if not category:
            continue
        counters["relevantProducts"] += 1
        shape = str(product.get("xingzhuang") or "").strip()
        package_key = "stacked" if category == stacked else package_by_shape.get(shape)
        if category == liquid and package_key in {"snapin", "screw", "t"}:
            category = large_liquid
        if not package_key:
            counters["missingPackage"] += 1
            unmatched.append(
                {
                    "itemNo": product.get("liaohao"),
                    "series": product.get("pinxing"),
                    "productCategory": category_titles.get(column_id, category),
                    "shape": shape,
                    "reason": "缺少可识别的封装类型",
                }
            )
            continue

        dimensions = (
            dimension_key(
                [product_number(product.get("slong")), product_number(product.get("swidth")), product_number(product.get("sheight"))]
            )
            if package_key == "stacked"
            else dimension_key([product_number(product.get("szhijing")), product_number(product.get("slong"))])
        )
        if not dimensions:
            counters["missingDimensions"] += 1
            unmatched.append(
                {
                    "itemNo": product.get("liaohao"),
                    "series": product.get("pinxing"),
                    "productCategory": category_titles.get(column_id, category),
                    "shape": shape,
                    "reason": "产品尺寸字段不完整",
                }
            )
            continue

        candidates = model_groups.get((category, package_key, dimensions), [])
        if not candidates:
            counters["noCadSize"] += 1
            unmatched.append(
                {
                    "itemNo": product.get("liaohao"),
                    "series": product.get("pinxing"),
                    "productCategory": category_titles.get(column_id, category),
                    "shape": shape,
                    "nominalDimensions": list(dimensions),
                    "reason": "当前 CAD 库没有同类别、同封装、同名义尺寸模型",
                }
            )
            continue

        # 一个哈希代表一个真实几何；重复路径不应制造多个产品候选。
        geometries: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for model in candidates:
            geometries[model["fileHash"]].append(model)
        item_token = match_token(product.get("liaohao"))
        series_token = match_token(product.get("pinxing"))

        def named_geometries(kind: str) -> set[str]:
            result: set[str] = set()
            for file_hash, hash_models in geometries.items():
                names = [match_token(model["fileName"]) for model in hash_models]
                if kind == "item" and item_token and any(item_token in name for name in names):
                    result.add(file_hash)
                elif kind == "series" and len(series_token) >= 3 and any(series_token in name for name in names):
                    result.add(file_hash)
                elif kind == "terminal":
                    suffix = "tm" if item_token.endswith("tm") else ("tg" if item_token.endswith("tg") else "")
                    if suffix and any(suffix in name for name in names):
                        result.add(file_hash)
            return result

        item_named = named_geometries("item")
        series_named = named_geometries("series")
        terminal_named = named_geometries("terminal")
        if len(item_named) == 1:
            selected_hashes = item_named
            match_method = "料号命中文件名"
            counters["uniqueByItemName"] += 1
        elif len(series_named) == 1:
            selected_hashes = series_named
            match_method = "系列命中文件名"
            counters["uniqueBySeriesName"] += 1
        elif len(terminal_named) == 1:
            selected_hashes = terminal_named
            match_method = "端子后缀命中文件名"
            counters["uniqueByTerminal"] += 1
        elif len(geometries) == 1:
            selected_hashes = set(geometries)
            match_method = "类别+封装+名义尺寸唯一"
            counters["uniqueByDimension"] += 1
        else:
            selected_hashes = set(geometries)
            match_method = "同尺寸存在多个几何候选"
            counters["multipleGeometry"] += 1

        ambiguous = len(selected_hashes) > 1
        status = "多候选待确认" if ambiguous else "唯一候选待确认"
        item_no = str(product.get("liaohao") or "").strip()
        series = str(product.get("pinxing") or "").strip()
        for file_hash in sorted(selected_hashes):
            hash_models = sorted(geometries[file_hash], key=lambda model: model["step"])
            model = hash_models[0]
            model_items[model["id"]].add(item_no)
            if series:
                model_series[model["id"]].add(series)
            if ambiguous:
                model_ambiguous[model["id"]].add(item_no)
            mappings.append(
                {
                    "databaseId": product.get("id"),
                    "itemNo": item_no,
                    "series": series,
                    "columnId": column_id,
                    "productCategory": category_titles.get(column_id, category),
                    "shape": shape,
                    "nominalDimensions": list(dimensions),
                    "cadModelId": model["id"],
                    "fileName": model["fileName"],
                    "step": model["step"],
                    "fileHash": model["fileHash"],
                    "category": model["category"],
                    "package": model["package"],
                    "subtype": model["subtype"],
                    "nominal": model["nominal"],
                    "cadEnvelope": model["cadEnvelope"],
                    "cadEnvelopeText": model["cadEnvelopeText"],
                    "candidatePaths": sum(len(geometries[value]) for value in selected_hashes),
                    "candidateGeometries": len(selected_hashes),
                    "match": match_method,
                    "status": status,
                }
            )

    for model in models:
        model["itemNos"] = sorted(model_items.get(model["id"], set()))
        model["series"] = sorted(model_series.get(model["id"], set()))
        model["productCandidateCount"] = len(model["itemNos"])
        model["productAmbiguousCount"] = len(model_ambiguous.get(model["id"], set()))

    mappings.sort(key=lambda row: (row["itemNo"], row["candidateGeometries"], row["cadModelId"]))
    summary = dict(counters)
    summary["mappedProducts"] = len({row["itemNo"] for row in mappings})
    summary["mappingRows"] = len(mappings)
    summary["unmatchedProducts"] = len(unmatched)
    summary["productionActiveRows"] = len(product_rows)
    coverage = {
        "source": "production:yongming.tp_good",
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "summary": summary,
        "unmatched": unmatched,
    }
    return mappings, coverage


def auxiliary_files() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in sorted((item for item in CAD_ROOT.rglob("*") if item.is_file() and item.suffix.lower() != ".stp"), key=lambda item: item.as_posix().lower()):
        name = path.name.lower()
        if "__out.log" in name:
            source_type = "Pro/ENGINEER 导出日志"
        elif ".asm." in name:
            source_type = "Creo/ProE 装配源文件"
        elif ".prt." in name:
            source_type = "Creo/ProE 零件源文件"
        else:
            source_type = "辅助源文件"
        relative = path.relative_to(CAD_ROOT)
        category, collection, _ = collection_info(relative)
        package, _, subtype = package_info(relative, category)
        rows.append(
            {
                "category": category,
                "collection": collection,
                "package": package,
                "subtype": subtype,
                "file": path.name,
                "extension": path.suffix,
                "sourceType": source_type,
                "fileSize": path.stat().st_size,
                "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                "path": path.relative_to(ROOT).as_posix(),
                "absolutePath": str(path),
                "pageUse": "不进入网页；保留作 CAD 源文件/日志",
            }
        )
    return rows


NAVY = "163A63"
BLUE = "2F75B5"
LIGHT_BLUE = "DDEBF7"
LIGHT_ORANGE = "FCE4D6"
LIGHT_GREEN = "E2F0D9"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="D9E2F0")


def style_title(sheet, title: str, subtitle: str, width: int) -> int:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = sheet.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=16)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    cell = sheet.cell(2, 1, subtitle)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.font = Font(color="44546A", size=10)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 26
    return 4


def style_data_sheet(sheet, headers: list[str], widths: list[int], table_name: str) -> None:
    header_row = 4
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=GRID)
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=GRID)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{sheet.cell(sheet.max_row, len(headers)).coordinate}"
    sheet.sheet_view.showGridLines = False
    if sheet.max_row >= 5:
        table = Table(displayName=table_name, ref=f"A4:{sheet.cell(sheet.max_row, len(headers)).coordinate}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)


def write_inventory(
    models: list[dict[str, Any]],
    product_map: list[dict[str, Any]],
    issues: list[dict[str, str]],
    auxiliary: list[dict[str, Any]],
    version: str,
) -> Path:
    inventory_path = ROOT / f"3DCAD模型清单_{version}.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    style_title(summary, f"3DCAD 模型清单 · {version}", f"生成时间：{generated}｜数据源：3dcad文件｜网页仅使用标准 STEP", 6)

    duplicate_groups = {model.get("duplicateGroup") for model in models if model.get("duplicateGroup")}
    duplicate_files = sum(1 for model in models if model.get("duplicateGroup"))
    stats = [
        ("标准 STEP 文件", len(models), "全部进入网页索引"),
        ("唯一几何内容", len({model["fileHash"] for model in models}), "按 SHA-256 去重统计"),
        ("已解析真实包络", sum(model["geometryStatus"] == "parsed" for model in models), "来自 STEP 网格的 X/Y/Z 包络"),
        ("完全重复组", len(duplicate_groups), f"涉及 {duplicate_files} 个文件，冗余副本 {duplicate_files - len(duplicate_groups)} 个"),
        ("需核对的问题记录", sum(issue["level"] == "需核对" for issue in issues), "详见“重复与问题”工作表"),
        ("已有料号候选关系", len(product_map), f"覆盖 {len({row['itemNo'] for row in product_map})} 个料号，均需业务确认"),
        ("辅助源文件/日志", len(auxiliary), "不进入网页预览"),
        ("STEP 总容量", round(sum(model["fileSize"] for model in models) / 1048576, 2), "MB"),
    ]
    for column, value in enumerate(["指标", "数量", "说明"], 1):
        summary.cell(4, column, value)
    for row in stats:
        summary.append(row)
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["C"].width = 48

    start = summary.max_row + 3
    summary.cell(start, 1, "按产品类别汇总")
    summary.cell(start, 1).font = Font(bold=True, color=NAVY, size=12)
    summary.append(["产品类别", "STEP 数量", "占比"])
    category_counts = Counter(model["category"] for model in models)
    for category in sorted(category_counts, key=lambda item: CATEGORY_ORDER.get(item, 99)):
        count = category_counts[category]
        summary.append([category, count, count / len(models)])
        summary.cell(summary.max_row, 3).number_format = "0.0%"

    start = summary.max_row + 3
    summary.cell(start, 1, "按封装类型汇总")
    summary.cell(start, 1).font = Font(bold=True, color=NAVY, size=12)
    summary.append(["封装类型", "STEP 数量", "占比"])
    package_counts = Counter(model["package"] for model in models)
    for package, count in sorted(package_counts.items(), key=lambda item: (-item[1], item[0])):
        summary.append([package, count, count / len(models)])
        summary.cell(summary.max_row, 3).number_format = "0.0%"
    summary.freeze_panes = "A4"
    summary.sheet_view.showGridLines = False

    model_sheet = workbook.create_sheet("模型清单")
    headers = ["序号", "产品类别", "封装类型", "结构/座板", "名义尺寸", "文件名尺寸", "CAD包络 X mm", "CAD包络 Y mm", "CAD包络 Z mm", "CAD部件数", "CAD三角面数", "CAD 文件名", "文件大小 KB", "修改时间", "相对路径", "SHA-256（前12位）", "重复组", "重复数量", "尺寸核对", "网页状态", "生产库候选系列", "生产库候选料号数", "其中多候选料号数", "制作表参考编号", "制作表系列", "引脚间距 mm"]
    style_title(model_sheet, f"标准 STEP 模型清单 · 共 {len(models)} 个", "CAD包络 X/Y/Z 来自 STEP 真实几何；名义尺寸用于封装检索。重复项未删除，已单独标记。", len(headers))
    for column, value in enumerate(headers, 1):
        model_sheet.cell(4, column, value)
    for index, model in enumerate(models, 1):
        model_sheet.append([
            index, model["category"], model["package"], model["subtype"], model["nominal"], model["actualSize"],
            *(model["cadEnvelope"] if len(model["cadEnvelope"]) == 3 else [None, None, None]),
            model["cadParts"], model["cadTriangles"], model["fileName"], round(model["fileSize"] / 1024, 1), model["modified"], model["step"], model["hashShort"], model.get("duplicateGroup", ""),
            model.get("duplicateCount", 1), "需核对" if model["dimensionMismatch"] else "正常", "已进入网页索引",
            "、".join(model["series"]), model["productCandidateCount"], model["productAmbiguousCount"],
            "、".join(model["sourceRefs"]), "、".join(model["sourceSeries"]), " / ".join(str(value) for value in model["pitches"]),
        ])
        path_cell = model_sheet.cell(model_sheet.max_row, 15)
        path_cell.hyperlink = Path(model["_absolutePath"]).as_uri()
        path_cell.style = "Hyperlink"
        if model["dimensionMismatch"]:
            model_sheet.cell(model_sheet.max_row, 19).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
        if model.get("duplicateGroup"):
            model_sheet.cell(model_sheet.max_row, 17).fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
    style_data_sheet(model_sheet, headers, [7, 18, 22, 14, 16, 16, 14, 14, 14, 12, 15, 40, 14, 20, 70, 18, 14, 12, 13, 16, 26, 16, 18, 30, 22, 16], "CadModelInventory")

    mapping_sheet = workbook.create_sheet("产品映射")
    mapping_headers = ["序号", "生产库ID", "料号", "产品系列", "产品类别", "产品形状", "产品名义尺寸", "CAD模型ID", "CAD文件名", "CAD类别", "CAD封装", "CAD结构", "CAD名义尺寸", "CAD真实包络", "匹配方式", "候选几何数", "确认状态", "STEP路径"]
    style_title(mapping_sheet, f"料号—CAD 候选映射 · 共 {len(product_map)} 条", "这是自动生成的候选关系。正式接入产品详情页前，必须按脚距、座板、端子结构和版本完成人工确认。", len(mapping_headers))
    for column, value in enumerate(mapping_headers, 1):
        mapping_sheet.cell(4, column, value)
    for index, row in enumerate(product_map, 1):
        mapping_sheet.append([
            index, row["databaseId"], row["itemNo"], row["series"], row["productCategory"], row["shape"],
            format_dimensions(row["nominalDimensions"]), row["cadModelId"], row["fileName"], row["category"], row["package"], row["subtype"], row["nominal"], row["cadEnvelopeText"], row["match"], row["candidateGeometries"], row["status"], row["step"],
        ])
        status_cell = mapping_sheet.cell(mapping_sheet.max_row, 17)
        status_cell.fill = PatternFill("solid", fgColor=LIGHT_ORANGE)
        path_cell = mapping_sheet.cell(mapping_sheet.max_row, 18)
        path_cell.hyperlink = (ROOT / Path(row["step"])).as_uri()
        path_cell.style = "Hyperlink"
    style_data_sheet(mapping_sheet, mapping_headers, [7, 12, 22, 14, 30, 16, 18, 18, 38, 18, 22, 14, 16, 30, 26, 14, 20, 72], "CadProductMap")

    issue_sheet = workbook.create_sheet("重复与问题")
    issue_headers = ["序号", "级别", "问题类型", "重复组", "产品类别", "封装", "名义尺寸", "文件名尺寸", "CAD 文件名", "相对路径", "说明", "建议"]
    style_title(issue_sheet, f"重复与问题 · 共 {len(issues)} 条记录", "“需核对”表示目录尺寸或归属可能有误；“提示”表示内容重复但不一定错误。源文件未做任何删除或移动。", len(issue_headers))
    for column, value in enumerate(issue_headers, 1):
        issue_sheet.cell(4, column, value)
    for index, issue in enumerate(issues, 1):
        issue_sheet.append([index, issue["level"], issue["type"], issue["group"], issue["category"], issue["package"], issue["nominal"], issue["actual"], issue["file"], issue["path"], issue["description"], issue["suggestion"]])
        fill = LIGHT_ORANGE if issue["level"] == "需核对" else LIGHT_GREEN
        issue_sheet.cell(issue_sheet.max_row, 2).fill = PatternFill("solid", fgColor=fill)
    style_data_sheet(issue_sheet, issue_headers, [7, 11, 28, 13, 18, 22, 16, 16, 38, 72, 42, 42], "CadIssues")

    aux_sheet = workbook.create_sheet("辅助源文件")
    aux_headers = ["序号", "产品类别", "封装", "结构", "文件类型", "扩展名/版本", "文件名", "文件大小 KB", "修改时间", "相对路径", "网页状态"]
    style_title(aux_sheet, f"辅助源文件与日志 · 共 {len(auxiliary)} 个", "这些文件不是标准 STEP；作为 Creo/ProE 源文件或导出日志保留，不进入浏览器预览和下载列表。", len(aux_headers))
    for column, value in enumerate(aux_headers, 1):
        aux_sheet.cell(4, column, value)
    for index, row in enumerate(auxiliary, 1):
        aux_sheet.append([index, row["category"], row["package"], row["subtype"], row["sourceType"], row["extension"], row["file"], round(row["fileSize"] / 1024, 1), row["modified"], row["path"], row["pageUse"]])
        path_cell = aux_sheet.cell(aux_sheet.max_row, 10)
        path_cell.hyperlink = Path(row["absolutePath"]).as_uri()
        path_cell.style = "Hyperlink"
    style_data_sheet(aux_sheet, aux_headers, [7, 18, 22, 14, 26, 14, 48, 14, 20, 76, 34], "CadAuxiliaryFiles")

    inventory_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(inventory_path)
    except PermissionError:
        date_stamp = datetime.now().strftime("%Y%m%d")
        inventory_path = ROOT / f"3DCAD模型清单_{version}_生产库更新_{date_stamp}.xlsx"
        workbook.save(inventory_path)
    return inventory_path


def public_model(model: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in model.items() if not key.startswith("_")}


def main() -> None:
    models, issues, version = build_models()
    auxiliary = auxiliary_files()
    product_map, product_coverage = build_product_map(models)
    public_models = [public_model(model) for model in models]
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(public_models, ensure_ascii=False, indent=2)
    OUTPUT.write_text(payload + "\n", encoding="utf-8")
    JS_OUTPUT.write_text("window.YMIN_CAD_MODELS = " + payload + ";\n", encoding="utf-8")
    PRODUCT_MAP_OUTPUT.write_text(json.dumps(product_map, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    PRODUCT_COVERAGE_OUTPUT.write_text(json.dumps(product_coverage, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    inventory_path = write_inventory(models, product_map, issues, auxiliary, version)

    category_counts = Counter(model["category"] for model in models)
    package_counts = Counter(model["package"] for model in models)
    duplicate_groups = len({model.get("duplicateGroup") for model in models if model.get("duplicateGroup")})
    duplicate_files = sum(1 for model in models if model.get("duplicateGroup"))
    print(f"版本：{version}")
    print(f"标准 STEP：{len(models)}，唯一几何：{len({model['fileHash'] for model in models})}")
    print("产品类别：" + "，".join(f"{key} {value}" for key, value in category_counts.items()))
    print("封装类型：" + "，".join(f"{key} {value}" for key, value in package_counts.items()))
    print(f"重复：{duplicate_groups} 组 / {duplicate_files} 个文件；辅助源文件：{len(auxiliary)}")
    print(f"问题记录：{len(issues)}，其中需核对 {sum(issue['level'] == '需核对' for issue in issues)}")
    print(f"真实包络：{sum(model['geometryStatus'] == 'parsed' for model in models)}/{len(models)}")
    print(f"生产库料号候选关系：{len(product_map)}，覆盖 {len({row['itemNo'] for row in product_map})} 个料号")
    print(f"已生成：{OUTPUT.relative_to(ROOT)}")
    print(f"已生成：{JS_OUTPUT.relative_to(ROOT)}")
    print(f"已生成：{PRODUCT_MAP_OUTPUT.relative_to(ROOT)}")
    print(f"已生成：{PRODUCT_COVERAGE_OUTPUT.relative_to(ROOT)}")
    print(f"已生成：{inventory_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
